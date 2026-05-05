from datetime import datetime, timezone
from pathlib import Path
from threading import Lock
from typing import Optional

import re
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

app = FastAPI()

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data" / "final"
FRONTEND_DIR = BASE_DIR / "frontend"
PREDICTIONS_FILE = DATA_DIR / "predictions_history.xlsx"
PREDICTION_LOCK = Lock()
PREDICTION_HEADERS = [
    "timestamp",
    "risk",
    "confidence",
    "risk_label",
    "fields_used_count",
    "total_fields",
    "age",
    "sex",
    "bmi",
    "cholesterol",
    "systolic_bp",
    "diastolic_bp",
    "glucose",
    "physical_activity",
    "smoking",
    "hypertension",
    "diabetes",
    "heart_rate",
    "ejection_fraction",
    "serum_creatinine",
    "source",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def resolve_risk_label(risk: float) -> str:
    if risk < 30:
        return "Low Risk"
    if risk < 60:
        return "Moderate Risk"
    return "High Risk"


FIELD_SPECS = {
    "age": {"weight": 1.2, "score": lambda value: clamp(((value - 18) / 62) * 100, 5, 95)},
    "sex": {"weight": 0.3, "score": lambda value: 56 if int(value) == 1 else 46},
    "bmi": {
        "weight": 1.0,
        "score": lambda value: 18 if 18.5 <= value < 25 else 48 if value < 30 else 74 if value < 35 else 88,
    },
    "cholesterol": {
        "weight": 0.9,
        "score": lambda value: 15 if value < 180 else 30 if value < 200 else 64 if value < 240 else 88,
    },
    "systolic_bp": {
        "weight": 1.2,
        "score": lambda value: 15 if value < 120 else 30 if value < 130 else 56 if value < 140 else 78 if value < 160 else 92,
    },
    "diastolic_bp": {
        "weight": 0.8,
        "score": lambda value: 15 if value < 80 else 45 if value < 90 else 72 if value < 100 else 88,
    },
    "glucose": {"weight": 1.0, "score": lambda value: 15 if value < 100 else 58 if value < 126 else 88},
    "physical_activity": {"weight": 0.8, "score": lambda value: clamp(90 - (value * 12), 12, 90)},
    "smoking": {"weight": 1.1, "score": lambda value: 84 if int(value) == 1 else 15},
    "hypertension": {"weight": 1.1, "score": lambda value: 86 if int(value) == 1 else 18},
    "diabetes": {"weight": 1.2, "score": lambda value: 88 if int(value) == 1 else 18},
    "heart_rate": {
        "weight": 0.6,
        "score": lambda value: 18 if 55 <= value <= 90 else 48 if 91 <= value <= 110 else 70,
    },
    "ejection_fraction": {
        "weight": 1.2,
        "score": lambda value: 15 if value >= 55 else 52 if value >= 41 else 78 if value >= 30 else 92,
    },
    "serum_creatinine": {
        "weight": 1.0,
        "score": lambda value: 18 if value < 1.2 else 58 if value < 2 else 78 if value < 3 else 92,
    },
}
TOTAL_WEIGHT = sum(spec["weight"] for spec in FIELD_SPECS.values())


def ensure_predictions_workbook() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if PREDICTIONS_FILE.exists():
        return

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Predictions"
    worksheet.append(PREDICTION_HEADERS)
    workbook.save(PREDICTIONS_FILE)


def read_predictions_history() -> list[dict]:
    ensure_predictions_workbook()

    with PREDICTION_LOCK:
        workbook = load_workbook(PREDICTIONS_FILE, data_only=True)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        items = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue
            items.append(dict(zip(headers, row)))

        return items


def append_prediction_record(record: dict) -> None:
    ensure_predictions_workbook()

    with PREDICTION_LOCK:
        workbook = load_workbook(PREDICTIONS_FILE)
        worksheet = workbook.active
        worksheet.append([record.get(header, "") for header in PREDICTION_HEADERS])
        workbook.save(PREDICTIONS_FILE)


class PatientData(BaseModel):
    age: Optional[float] = None
    sex: Optional[int] = None
    bmi: Optional[float] = None
    cholesterol: Optional[float] = None
    systolic_bp: Optional[float] = None
    diastolic_bp: Optional[float] = None
    glucose: Optional[float] = None
    smoking: Optional[int] = None
    hypertension: Optional[int] = None
    physical_activity: Optional[float] = None
    diabetes: Optional[int] = None
    heart_rate: Optional[float] = None
    ejection_fraction: Optional[float] = None
    serum_creatinine: Optional[float] = None


def sanitize_payload(data: PatientData) -> dict:
    raw_payload = data.model_dump(exclude_none=True)
    payload = {}

    for key, value in raw_payload.items():
        if key not in FIELD_SPECS:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        payload[key] = value

    return payload


def calculate_prediction(payload: dict) -> tuple[float, float, dict]:
    weighted_total = 0.0
    active_weight = 0.0
    breakdown = {}

    for field_name, value in payload.items():
        spec = FIELD_SPECS.get(field_name)
        if not spec:
            continue

        field_score = float(spec["score"](value))
        field_weight = float(spec["weight"])
        weighted_total += field_score * field_weight
        active_weight += field_weight
        breakdown[f"{field_name}_factor"] = round(field_score, 2)

    if active_weight == 0:
        raise HTTPException(status_code=400, detail="Please enter at least 2 values to predict risk.")

    risk = round(weighted_total / active_weight, 2)
    completeness = active_weight / TOTAL_WEIGHT
    confidence = round(clamp(58 + (completeness * 30) + min(abs(risk - 50) / 12, 6), 58, 94), 2)
    return risk, confidence, breakdown


def build_prediction_record(payload: dict, risk: float, confidence: float) -> dict:
    record = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "risk": round(risk, 2),
        "confidence": round(confidence, 2),
        "risk_label": resolve_risk_label(risk),
        "fields_used_count": len(payload),
        "total_fields": len(FIELD_SPECS),
        "source": "backend",
    }

    for header in PREDICTION_HEADERS:
        if header in payload:
            record[header] = payload[header]
        elif header not in record:
            record[header] = ""

    return record


@app.get("/api/health")
def health():
    return {"message": "HeartGuard API Running"}


@app.get("/predictions")
def get_predictions():
    try:
        return {"items": read_predictions_history()}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not load prediction history: {exc}") from exc


@app.post("/predict")
def predict(data: PatientData):
    try:
        payload = sanitize_payload(data)

        if len(payload) < 2:
            raise HTTPException(status_code=400, detail="Please enter at least 2 values to predict risk.")

        risk, confidence, breakdown = calculate_prediction(payload)
        record = build_prediction_record(payload, risk, confidence)
        append_prediction_record(record)

        return {
            "final_risk_percentage": round(risk, 2),
            "confidence": round(confidence, 2),
            "risk_label": record["risk_label"],
            "fields_used_count": record["fields_used_count"],
            "total_fields": record["total_fields"],
            "saved_record": record,
            "model_breakdown": breakdown,
        }

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


class ChatRequest(BaseModel):
    message: str


@app.post("/chat")
def chat(req: ChatRequest):
    try:
        user_msg = req.message.strip().lower()
        reply = generate_heartguard_reply(user_msg)
        return {"reply": reply}

    except Exception as exc:
        return {"reply": f"Error: {str(exc)}"}


def generate_heartguard_reply(user_msg: str) -> str:
    if not user_msg:
        return (
            "Please type a question about heart health, diet, exercise, blood pressure, "
            "cholesterol, diabetes, or lifestyle habits."
        )

    if any(word in user_msg for word in ["hello", "hi", "hey", "good morning", "good evening"]):
        return (
            "Hello! I am your HeartGuard AI Health Assistant.\n\n"
            "I can help you with:\n"
            "- heart-health symptoms and warning signs\n"
            "- blood pressure, cholesterol, and diabetes questions\n"
            "- diet, exercise, sleep, stress, and smoking guidance\n"
            "- practical lifestyle steps for reducing heart risk\n\n"
            "Ask me anything specific, for example: "
            "\"How can I lower my cholesterol naturally?\""
        )

    patterns = [
        (r"early signs|symptoms|heart disease|heart attack warning", answer_early_signs),
        (r"cholesterol|ldl|hdl|triglyceride", answer_cholesterol),
        (r"blood pressure|bp|hypertension", answer_blood_pressure),
        (r"exercise|workout|walking|cardio|strength training", answer_exercise),
        (r"diet|food|meal|breakfast|eat|nutrition", answer_diet),
        (r"smoking|quit smoking|cigarette|tobacco", answer_smoking),
        (r"stress|anxiety|breathing|relax", answer_stress),
        (r"sleep|sleep duration|insomnia|rest", answer_sleep),
        (r"diabetes|sugar|glucose|blood sugar", answer_diabetes),
        (r"weight|bmi|obesity", answer_weight),
    ]

    for pattern, handler in patterns:
        if re.search(pattern, user_msg):
            return handler(user_msg)

    return (
        "Heart health usually improves when you focus on a few core habits consistently:\n\n"
        "1. Keep blood pressure, cholesterol, blood sugar, and weight in a healthy range.\n"
        "2. Exercise regularly, even if it starts with brisk walking.\n"
        "3. Eat more whole foods like vegetables, fruit, legumes, oats, nuts, and fish.\n"
        "4. Avoid smoking and limit highly processed, salty, or sugary foods.\n"
        "5. Get enough sleep and manage stress in a sustainable way.\n\n"
        "If you want a more detailed answer, ask a narrower question such as:\n"
        "- What are the early signs of heart disease?\n"
        "- How can I lower cholesterol naturally?\n"
        "- What foods should I avoid for heart health?\n"
        "- What exercise is safest if I am overweight?"
    )


def answer_early_signs(_: str) -> str:
    return (
        "Early signs of heart disease can be subtle, and some people have no symptoms at first. "
        "Common warning signs include:\n\n"
        "- chest pain, pressure, tightness, or burning\n"
        "- shortness of breath during activity or even at rest\n"
        "- unusual fatigue, especially with routine tasks\n"
        "- dizziness, fainting, or reduced exercise tolerance\n"
        "- palpitations or an irregular heartbeat\n"
        "- swelling in the legs, ankles, or feet\n"
        "- pain spreading to the arm, jaw, neck, shoulder, or back\n\n"
        "Important: women, older adults, and people with diabetes may have less obvious symptoms, "
        "such as fatigue, nausea, indigestion-like discomfort, or breathlessness.\n\n"
        "Seek urgent medical help immediately if chest pain is severe, lasts more than a few minutes, "
        "or is combined with sweating, nausea, fainting, or shortness of breath."
    )


def answer_cholesterol(_: str) -> str:
    return (
        "To lower cholesterol naturally, focus on habits that reduce LDL and improve overall heart health:\n\n"
        "- eat more soluble fiber from oats, beans, lentils, fruits, and vegetables\n"
        "- replace fried and processed foods with nuts, seeds, olive oil, and fish\n"
        "- reduce saturated fat from fatty meats, processed snacks, and excess butter or ghee\n"
        "- avoid trans fats entirely when possible\n"
        "- exercise regularly, aiming for at least 150 minutes per week\n"
        "- maintain a healthy weight if you are overweight\n"
        "- stop smoking and limit alcohol\n\n"
        "As a general reference, total cholesterol is often preferred below 200 mg/dL, "
        "but LDL, HDL, and triglycerides matter too. If your LDL is high or you already have heart disease, "
        "your doctor may recommend stricter targets and sometimes medication."
    )


def answer_blood_pressure(_: str) -> str:
    return (
        "Blood pressure matters because long-term elevation puts extra strain on the heart, blood vessels, kidneys, and brain.\n\n"
        "General reference values:\n"
        "- around 120/80 mmHg is often considered healthy\n"
        "- consistently elevated readings deserve follow-up\n"
        "- very high readings, especially with headache, chest pain, confusion, or breathlessness, need urgent care\n\n"
        "Ways to improve blood pressure:\n"
        "- reduce excess salt and highly processed foods\n"
        "- stay active most days of the week\n"
        "- sleep well and manage chronic stress\n"
        "- maintain a healthy weight\n"
        "- stop smoking and moderate alcohol intake\n"
        "- take prescribed medicines consistently if your doctor has started treatment\n\n"
        "The most useful approach is regular home monitoring plus medical follow-up."
    )


def answer_exercise(user_msg: str) -> str:
    if "best" in user_msg or "heart health" in user_msg:
        intro = "The best exercise plan for heart health usually combines aerobic activity with strength work.\n\n"
    else:
        intro = "Exercise is one of the most effective ways to reduce cardiovascular risk.\n\n"

    return (
        intro
        + "A balanced weekly routine often includes:\n"
        "- brisk walking, cycling, swimming, or light jogging for aerobic fitness\n"
        "- strength training 2 to 3 times per week\n"
        "- mobility or stretching for consistency and recovery\n\n"
        "A practical target is at least 150 minutes of moderate activity per week. "
        "If you are just starting, even 10 to 15 minutes per day is a good beginning.\n\n"
        "Stop and seek medical advice if exercise causes chest pain, severe breathlessness, dizziness, or fainting."
    )


def answer_diet(user_msg: str) -> str:
    if "breakfast" in user_msg:
        return (
            "A heart-healthy breakfast should include fiber, protein, and minimally processed foods.\n\n"
            "Good options include:\n"
            "- oatmeal with fruit, chia seeds, and a few nuts\n"
            "- plain yogurt or curd with fruit and seeds\n"
            "- vegetable omelet with whole-grain toast\n"
            "- poha, upma, or dal chilla prepared with less oil and more vegetables\n"
            "- smoothie with unsweetened milk, oats, fruit, and seeds\n\n"
            "Try to avoid sugary cereals, deep-fried breakfast foods, and very salty processed meats."
        )

    if "avoid" in user_msg:
        return (
            "For better heart health, try to limit foods that raise blood pressure, cholesterol, or inflammation:\n\n"
            "- deep-fried foods\n"
            "- processed meats\n"
            "- packaged snacks high in salt\n"
            "- sugary drinks and desserts in excess\n"
            "- bakery foods made with unhealthy fats\n"
            "- very large portions of refined carbohydrates\n\n"
            "Instead, build meals around vegetables, fruits, legumes, whole grains, nuts, seeds, and lean proteins."
        )

    return (
        "A heart-healthy diet is usually built on whole, minimally processed foods.\n\n"
        "A simple formula is:\n"
        "- half the plate vegetables\n"
        "- one quarter lean protein such as beans, lentils, fish, eggs, or chicken\n"
        "- one quarter whole grains or other high-fiber carbohydrates\n"
        "- healthy fats in moderate portions, such as nuts, seeds, avocado, or olive oil\n\n"
        "This kind of eating pattern can support blood pressure, cholesterol, blood sugar, and weight control."
    )


def answer_smoking(_: str) -> str:
    return (
        "Smoking significantly increases the risk of heart disease, stroke, and poor circulation because it damages blood vessels, "
        "raises clot risk, and reduces oxygen delivery.\n\n"
        "If you want to quit, these steps help most:\n"
        "- pick a quit date\n"
        "- identify triggers such as stress, tea breaks, or social cues\n"
        "- remove cigarettes and related items from your environment\n"
        "- ask your doctor about nicotine replacement or medicines if needed\n"
        "- tell family or friends so they can support you\n"
        "- expect cravings and plan alternatives like walking, water, chewing gum, or breathing exercises\n\n"
        "Even quitting later in life still gives meaningful cardiovascular benefit."
    )


def answer_stress(_: str) -> str:
    return (
        "Stress can affect heart health both directly and indirectly. It may raise blood pressure temporarily, "
        "worsen sleep, increase emotional eating, and make smoking or inactivity more likely.\n\n"
        "Helpful stress-management habits include:\n"
        "- regular physical activity\n"
        "- consistent sleep schedule\n"
        "- slow breathing or mindfulness practice\n"
        "- limiting constant screen or news overload\n"
        "- speaking with trusted people or a mental health professional when stress is persistent\n\n"
        "If stress is causing chest pain, panic-like symptoms, or worsening blood pressure, it deserves medical attention."
    )


def answer_sleep(_: str) -> str:
    return (
        "Sleep is an important part of heart health. For most adults, about 7 to 9 hours per night is a useful target.\n\n"
        "Poor sleep is linked with higher blood pressure, weight gain, insulin resistance, and increased cardiovascular risk.\n\n"
        "To improve sleep quality:\n"
        "- keep a regular bedtime and wake time\n"
        "- reduce caffeine late in the day\n"
        "- avoid heavy meals right before bed\n"
        "- keep the room dark and quiet\n"
        "- limit phone or laptop use before sleep\n\n"
        "If you snore heavily, wake up tired, or stop breathing during sleep, ask about sleep apnea evaluation."
    )


def answer_diabetes(_: str) -> str:
    return (
        "Diabetes and high blood sugar increase heart risk because they can damage blood vessels over time.\n\n"
        "The main goals are:\n"
        "- keep blood sugar in your target range\n"
        "- control blood pressure and cholesterol too, not just glucose\n"
        "- stay physically active\n"
        "- eat a balanced, lower-refined-carb diet\n"
        "- attend regular medical follow-up and medication review\n\n"
        "People with diabetes may also have less typical heart symptoms, so unexplained fatigue or breathlessness should not be ignored."
    )


def answer_weight(_: str) -> str:
    return (
        "Excess body weight, especially around the abdomen, can increase the risk of high blood pressure, diabetes, abnormal cholesterol, "
        "and heart disease.\n\n"
        "A useful approach is gradual, sustainable change:\n"
        "- improve meal quality rather than crash dieting\n"
        "- increase walking and weekly activity\n"
        "- prioritize sleep and stress control\n"
        "- watch portion size and liquid calories\n\n"
        "Even modest weight loss can improve cardiovascular markers if it is maintained."
    )


@app.get("/")
def serve_home():
    return FileResponse(FRONTEND_DIR / "index.html")


@app.get("/{requested_path:path}")
def serve_frontend_file(requested_path: str):
    normalized_path = requested_path.strip("/")

    if not normalized_path:
        return FileResponse(FRONTEND_DIR / "index.html")

    safe_path = (FRONTEND_DIR / normalized_path).resolve()

    if safe_path.is_file() and FRONTEND_DIR.resolve() in safe_path.parents:
        return FileResponse(safe_path)

    raise HTTPException(status_code=404, detail="Not Found")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="127.0.0.1", port=8000)
